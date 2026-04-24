"""Ingest Urząd Statystyczny Excel files → data/processed/urzad_meta.json.

Reads each Excel/XLS file in the Urząd directory and produces a metadata
JSON listing: file name, sheet names, approximate row count per sheet,
and sampled column headers.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from src.config import URZAD_DIR, URZAD_META_JSON


def _inspect_xlsx(path: Path) -> dict:
    """Return metadata dict for one XLSX file."""
    entry: dict = {"file": path.name, "sheets": []}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            # Find first non-empty row as header sample
            header_sample: list[str] = []
            for row in all_rows[:10]:
                non_null = [str(v) for v in row if v is not None]
                if non_null:
                    header_sample = non_null[:8]
                    break
            entry["sheets"].append({
                "name":          sheet_name,
                "approx_rows":   len(all_rows),
                "header_sample": header_sample,
            })
        wb.close()
    except Exception as exc:
        entry["error"] = str(exc)
    return entry


def _inspect_xls(path: Path) -> dict:
    """Return metadata dict for one legacy XLS file (via xlrd)."""
    entry: dict = {"file": path.name, "sheets": []}
    try:
        import xlrd  # type: ignore
        wb = xlrd.open_workbook(str(path))
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            header_sample: list[str] = []
            for r in range(min(10, ws.nrows)):
                non_null = [str(ws.cell_value(r, c)) for c in range(ws.ncols)
                            if ws.cell_value(r, c) not in ("", None)]
                if non_null:
                    header_sample = non_null[:8]
                    break
            entry["sheets"].append({
                "name":          sheet_name,
                "approx_rows":   ws.nrows,
                "header_sample": header_sample,
            })
    except Exception as exc:
        entry["error"] = str(exc)
    return entry


def load(urzad_dir: Path = URZAD_DIR) -> list[dict]:
    """Inspect all Excel files and return a list of metadata dicts."""
    meta: list[dict] = []
    for path in sorted(urzad_dir.glob("*")):
        if path.suffix.lower() == ".xlsx":
            meta.append(_inspect_xlsx(path))
        elif path.suffix.lower() == ".xls":
            meta.append(_inspect_xls(path))
    return meta


def main() -> None:
    """Inspect files, save metadata JSON, and print a summary."""
    meta = load()
    URZAD_META_JSON.parent.mkdir(parents=True, exist_ok=True)
    with URZAD_META_JSON.open("w", encoding="utf-8") as fh:
        json.dump(meta, fh, ensure_ascii=False, indent=2)

    print(f"Saved → {URZAD_META_JSON}")
    for entry in meta:
        print(f"\n  {entry['file']}")
        if "error" in entry:
            print(f"    ERROR: {entry['error']}")
            continue
        for sheet in entry["sheets"]:
            print(f"    sheet '{sheet['name']}': ~{sheet['approx_rows']} rows")
            print(f"      header: {sheet['header_sample']}")


if __name__ == "__main__":
    main()
