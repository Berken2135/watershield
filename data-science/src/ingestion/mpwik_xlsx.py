"""Ingest MPWiK sample-collection XLSX files → data/processed/mpwik_samples.parquet.

Each XLSX has the structure:
  Row 0: [None, None, param_pl_1, param_pl_2, …]
  Row 1: ['Punkt adresowy', 'Data/godzina pobierania', unit_1, unit_2, …]
  Row 2+: [location_or_None, timestamp, val_1, val_2, …]

Location is forward-filled from the first non-null value in column 0.
Output long format: timestamp, location, parameter, value, unit.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Iterator

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from src.config import MPWIK_SAMPLES_PARQUET, MPWIK_XLSX_DIR


def _parse_xlsx(path: Path) -> Iterator[dict]:
    """Yield long-format row dicts from a single sample-collection XLSX."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        print(f"  [skip] {path.name}: {exc}")
        return

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        param_row = rows[0]   # Polish parameter names (cols 2+)
        unit_row  = rows[1]   # units (cols 2+)
        data_rows = rows[2:]

        # Build list of (col_index, parameter_name, unit) for value columns
        value_cols: list[tuple[int, str, str]] = []
        for col_idx in range(2, len(param_row)):
            param = param_row[col_idx]
            unit  = unit_row[col_idx] if col_idx < len(unit_row) else None
            if param is not None:
                value_cols.append((col_idx, str(param).strip(), str(unit).strip() if unit else ""))

        current_location: str | None = None
        for row in data_rows:
            # Forward-fill location
            loc_cell = row[0] if len(row) > 0 else None
            if loc_cell is not None and str(loc_cell).strip():
                current_location = str(loc_cell).strip()

            ts_cell = row[1] if len(row) > 1 else None
            if ts_cell is None:
                continue
            try:
                timestamp = pd.Timestamp(ts_cell)
            except (ValueError, TypeError):
                continue

            for col_idx, parameter, unit in value_cols:
                val = row[col_idx] if col_idx < len(row) else None
                if val is None or str(val).strip() == "":
                    continue
                try:
                    value = float(str(val).replace(",", "."))
                except (ValueError, TypeError):
                    continue

                yield {
                    "timestamp": timestamp,
                    "location":  current_location or path.stem,
                    "parameter": parameter,
                    "value":     value,
                    "unit":      unit,
                }


def load(xlsx_dir: Path = MPWIK_XLSX_DIR) -> pd.DataFrame:
    """Read all sample XLSX files (except Legend) and return long-format DataFrame."""
    rows: list[dict] = []
    for path in sorted(xlsx_dir.glob("*.xlsx")):
        if path.stem.lower() == "legend":
            continue
        rows.extend(_parse_xlsx(path))

    if not rows:
        raise ValueError(f"No data parsed from {xlsx_dir}")

    df = pd.DataFrame(rows)
    df = df.sort_values(["location", "parameter", "timestamp"]).reset_index(drop=True)
    return df


def main() -> None:
    """Load, save as parquet, and print a summary."""
    df = load()
    MPWIK_SAMPLES_PARQUET.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(MPWIK_SAMPLES_PARQUET, index=False)

    print(f"Saved → {MPWIK_SAMPLES_PARQUET}")
    print(f"  rows      : {len(df):,}")
    print(f"  date range: {df['timestamp'].min()}  →  {df['timestamp'].max()}")
    print(f"  locations : {sorted(df['location'].unique())}")
    print(f"  parameters: {sorted(df['parameter'].unique())}")
    null_counts = df.isnull().sum()
    nonzero = null_counts[null_counts > 0]
    if nonzero.empty:
        print("  nulls     : none")
    else:
        print("  nulls     :")
        for col, n in nonzero.items():
            print(f"    {col}: {n}")


if __name__ == "__main__":
    main()
